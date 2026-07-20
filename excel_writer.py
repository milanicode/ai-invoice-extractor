from pathlib import Path

from openpyxl import Workbook, load_workbook

COLUMNS = [
    "source_file",
    "provider",
    "vendor_name",
    "vendor_tax_id",
    "invoice_number",
    "issue_date",
    "due_date",
    "currency",
    "subtotal",
    "tax",
    "total",
    "notes",
]


def _row(source_file: str, provider: str, data: dict) -> list:
    return [
        source_file,
        provider,
        data.get("vendor_name"),
        data.get("vendor_tax_id"),
        data.get("invoice_number"),
        data.get("issue_date"),
        data.get("due_date"),
        data.get("currency"),
        data.get("subtotal"),
        data.get("tax"),
        data.get("total"),
        data.get("notes"),
    ]


def reset_workbook(xlsx_path: Path) -> None:
    """Create a new Excel file with headers only (overwrite if it exists)."""
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "invoices"
    ws.append(COLUMNS)
    wb.save(xlsx_path)


def append_invoice(xlsx_path: Path, source_file: str, provider: str, data: dict) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "invoices"
        ws.append(COLUMNS)

    ws.append(_row(source_file, provider, data))
    wb.save(xlsx_path)
